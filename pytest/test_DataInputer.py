import openpyxl
import os
import time

def get_base_path():
    """기본 경로를 반환합니다."""
    return os.path.dirname(os.path.abspath(os.path.dirname(__file__)))

def load_workbook_data(file_path):
    """엑셀 파일을 열고 데이터만 가져옵니다."""
    return openpyxl.load_workbook(file_path, data_only=True)

def create_slk_header():
    """SLK 헤더를 생성합니다."""
    return ["ID;P"]

def get_col_mappings():
    """열별 매핑 데이터를 반환합니다."""
    return {
        2: { '이치고': 1, '루키아': 2, '우류': 3, '렌지': 4, '오리히메': 5, '사도': 6 },
        3: { '기본': 1, '핵심': 2, '핵심 [必]': 3, '패시브': 4, '변신': 5 },
        4: { '회피': 1, '보법': 2, '점프': 3, '횡베기': 4, '찌르기': 5, '빔': 6,
             '종베기': 7, '전투': 8, '은통': 9, '영자병장': 10, '귀도': 11, '검무': 12,
             '공격': 13, '꽃': 14, '조합 사용': 15, '지속 스킬': 16, '참술': 17,
             '백타': 18, '주먹': 19, '대지강타': 20, '방어': 21, '변신': 22,
             '특성 강화': 23, '변신 강화': 24, '스킬 강화': 25, '온도제어': 26 },
        6: { '이동기': 1, '강화버프': 2, '사출기': 3, '스위치오라': 4, '돌진기': 5,
             '영구지속': 6, '타겟팅': 7, '주변범위': 8, '광역기' : 9 },
        7: { '클릭불가': 1, '즉발': 2, '대상형': 3, '방향형': 4, '범위형': 5 }

    }



def process_sheet_str(input_col, input_type, input_value):
    solo_data = 'call SaveStr(udg_hash, StringHash("' + keyword + '"), ' + str(row[0]) + ', "' + row[7] + '//' + row[8] + '")'
        
    str_data = []
    slk_content.append(f"C;X{col_idx};Y{row_idx-start_row + 1};K\"{write_value}\"")


def save_str_data(keyword, data):
    with open(os.path.join(get_base_path(), 'Data', keyword + '.j'), 'w', encoding='utf-8') as result:
        result.write('library ' + keyword + ' initializer Init\n')
        result.write('// 추가한 시간: ' + time.strftime('%y.%m.%d %X') + '\n')
        result.write('\n  private function Init takes nothing returns nothing\n')
        for child in data:
            result.write('    ' + child+'\n')
        result.write('  endfunction\nendlibrary')
        print(f'completed {keyword}[{len(data)}]')
        
        update_git_adder(get_base_path(), keyword + '.j')
        
def process_sheet(sheet, start_row, end_col, col_mappings, specific_rows, input_keyword):
    """시트를 처리하여 SLK 내용을 생성합니다."""
    slk_content = create_slk_header()
    str_content = []

    for row_idx in range(start_row, sheet.max_row + 1):
        if not should_process_row(sheet, row_idx, start_row):
            break  # 1번째 열이 없으면 저장 종료

        row = sheet[row_idx]
        for col_idx in range(1, end_col + 1):
            write_value = row[col_idx - 1].value  # 0-based indexing

            if row_idx == start_row:
                add_slk_header(slk_content, col_idx, write_value, row_idx, start_row)
                continue

            if col_idx in specific_rows:
                add_specific_row(str_content, input_keyword, sheet, write_value, col_idx, row_idx, start_row)
                continue

            if not is_value_storable(col_idx, write_value):
                continue

            write_value = map_write_value(write_value, col_mappings, col_idx)
            slk_content.append(f"C;X{col_idx};Y{row_idx - start_row + 1};K{write_value}")

    save_str_data(f"{input_keyword}Str", str_content)
    return slk_content

def should_process_row(sheet, row_idx, start_row):
    """행을 처리할 수 있는지 여부를 판단합니다."""
    return row_idx == start_row or (sheet.cell(row_idx, column=1).value != 0 and sheet.cell(row_idx, column=1).value is not None)

def add_slk_header(slk_content, col_idx, write_value, row_idx, start_row):
    """SLK 헤더를 추가합니다."""
    slk_content.append(f"C;X{col_idx};Y{row_idx - start_row + 1};K\"{write_value}\"")

def add_specific_row(str_content, input_keyword, sheet, write_value, col_idx, row_idx, start_row):
    """특정 행의 데이터를 추가합니다."""
    str_content.append(f'set s__{input_keyword}_{sheet.cell(start_row, col_idx).value}[{row_idx - start_row}]="{write_value}"')

def is_value_storable(col_idx, write_value):
    """저장할 수 있는 값인지 여부를 판단합니다."""
    return not (col_idx >= 7 and (write_value is None or write_value == 0))

def map_write_value(write_value, col_mappings, col_idx):
    """값을 매핑합니다."""
    if isinstance(write_value, str):
        mapping_data = col_mappings.get(col_idx, {})
        return mapping_data.get(write_value, f'"{write_value}"')
    return write_value

def save_file(file_path, file_content):
    """파일을 저장합니다."""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(file_content) + '\n')

def update_git_adder(base_path, keyword):
    """GitAdder 파일을 업데이트합니다."""
    git_adder_path = os.path.join(base_path, 'GitAdder.j')
    
    with open(git_adder_path, 'r', encoding='utf-8') as adder:
        lines = adder.readlines()
    
    has_keyword = any(keyword in line for line in lines)
    
    if not has_keyword:
        with open(git_adder_path, 'a', encoding='utf-8') as adder:
            if ( base_path.__contains__('.slk')):
                adder.write(f'\n//! loaddata "Data\\{keyword}"')
            else:
                adder.write(f'\n//! import "Data\\{keyword}"')
        print('추가완료')
    else:
        print(f'{keyword} : 이미 추가되어 있음')

def main():
    base_path = get_base_path()
    excel_path = os.path.join(base_path, 'OMO RPG.xlsx')
    main_keyword = 'SkillData'
    slk_path = os.path.join(base_path, 'Data', main_keyword + '.slk')

    workbook = load_workbook_data(excel_path)
    sheet = workbook['스킬 데이터']
    
    col_mappings = get_col_mappings()
    slk_content = process_sheet(sheet, 7, 25, col_mappings, [8, 9], main_keyword)
    
    save_file(slk_path, slk_content)
    print(f"SLK 파일로 저장 완료[{len(slk_content)}]: {slk_path}")

    update_git_adder(base_path, main_keyword + '.slk')
    workbook.close()

if __name__ == '__main__':
    main()

