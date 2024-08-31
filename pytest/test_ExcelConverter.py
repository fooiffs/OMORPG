import openpyxl
import os
import time

# 상위 경로
base_path = os.path.dirname(os.path.abspath(os.path.dirname(__file__)))
# 2단 상위 경로
# os.path.dirname(os.path.abspath(os.path.dirname(os.path.abspath(os.path.dirname(__file__)))))

# 새로운 워크북 생성
workbook = openpyxl.load_workbook(base_path + '\\OMO RPG.xlsx')

# 엑셀 파일 열기
sheet = workbook['스킬 데이터']
 
# 저장 대상
keyword = 'SkillObject'

# 행 데이터 반복 처리
data = []
# print('max: ' + str(sheet.max_row))
for row in sheet.iter_rows(min_row=8, max_row=sheet.max_row, min_col=0, max_col=9, values_only = True):
    if ( type(row[0]) == int ): 
        solo_data = 'call SaveStr(udg_hash, StringHash("' + keyword + '"), ' + str(row[0]) + ', "' + row[7] + '//' + row[8] + '")'
        data.append(solo_data)
        # print(solo_data)
    else:
         # 파일 닫기
        workbook.close()

# # 새 파일 경로
txt_file_path = base_path + '\\Data\\' + keyword + '.j'

# txt 파일 쓰기
# print(data)
with open(txt_file_path, 'w', encoding='utf-8') as result:
    result.write('library ' + keyword + ' initializer init\n')
    result.write('// 추가한 시간: ' + time.strftime('%y.%m.%d %X') + '\n')
    result.write('\n  private function init takes nothing returns nothing\n')

    for child in data:
        result.write('    ' + child+'\n')
    result.write('  endfunction\nendlibrary')

print('completed [' + str(len(data))+ ']')


adder = open(base_path + '\\GitAdder.j', 'r')
has = False
data2 = []
for line in enumerate(adder):
    # 문자열 검색
    data2.append(line)

    if keyword in line[1]:
        has = True

print('completed [' + str(len(data2))+ ']')
if ( has == False ):
    # adder.close()
    adder = open(base_path + '\\GitAdder.j', 'w')

    for child2 in data2:
        adder.write(str(child2[1]))

    adder.write('\n' + '//! import "Data\\' + keyword + '.j"')

    # adder.close()
    print('추가완료')
else:
    print('이미 추가되어 있음')

